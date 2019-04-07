import os
import time
from datetime import datetime
from tkinter import Tk, messagebox

import numpy as np
import openpyxl as excel
import pandas as pd
import xlrd
import configparser

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select, WebDriverWait
from seleniumOperationBase import SeleniumOperationBase
from utils.logger import LoggerObj
from webBase import WebExecuteBase
from openpyxl.styles.borders import Border, Side
import openpyxl as excel
from openpyxl.utils import get_column_letter

driver=webdriver.Chrome('C:/webdrivers/chromedriver.exe')
#driver=webdriver.Firefox('C:/webdrivers/geckodriver.exe')
#driver=webdriver.Firefox()

# ログイン情報などを取得
iniFile=configparser.ConfigParser()



RESERVE_DATE='//*[@id="datePick"]'
REST_DAY='//*[@id="reserve_term"]'

NUMBER_PEOPLE='//*[@id="headcount"]'

NEED_MORNING='//div/div/form/input[{0}]'

PLAN_SELECT='//div/div/form/input[{0}]'

NAME='//*[@id="guestname"]'

MORNING_TYPE={'あり':5,'なし':6}
PLAN_TYPE={'昼からチェックインプラン':7,'お得な観光プラン':8}

NEXT_BUTTON='//*[@id="agree_and_goto_next"]'

PAY_DETAIL='/html/body/div[1]/div/form/div/div/a'
BACK_BUTTON='//div/div/button'
CONFIRM_BUTTON='//div/div/form/button'  
RESULT_BACK_BUTTON='//div/div/button'

CONFIRM_COST='//*[@id="price"]'
CONFIRM_DATE_FROM='//*[@id="datefrom"]'
CONFIRM_DATE_TO='//*[@id="dateto"]'
CONFIRM_DAYSCOUNT='//*[@id="dayscount"]'
CONFIRM_PEOPLE_COUNT='//*[@id="hc"]'
CONFIRM_NEED_MORNING='//*[@id="bf_order"]'
CONFIRM_PLAN='//h4[4]/span'
#//*[@id="plan_a_order"]'
CONFIRM_NAME='//*[@id="gname"]'

COLUMN_INFO=['項番','番号_宿泊者','料金','期間開始','期間終了','宿泊日数','宿泊人数','朝食有無','プラン','お名前']
COLUMN_SIZE_INFO=[0,30,10,15,15,14,14,14,22,20]


class TestSiteOrder(SeleniumOperationBase):

    def __init__(self,driver,log,screenShotBaseName='screenShotName'):
        super().__init__(driver,log,screenShotBaseName)
    
    def pullDownSelect(self,webElement,inputText):
        target=self.driver.find_element_by_xpath(webElement)
        selecttargetForm=Select(target)
        # 画面に表示されるプルダウンのテキストで選択を行う
        selecttargetForm.select_by_visible_text(inputText.zfill(2))

    def inputOrder(self,reserveSheet):
        reserveSheetDict=reserveSheet.to_dict('index')

        # 予約者名:[予約情報]の形
        restInfos={}
        bfPlan=''
        for index,data in reserveSheetDict.items():

            number=data['項番']
            if number !=number:
                # 項番がない状態であれば登録処理を終了
                break
            day=data['宿泊日'].replace('-','/')
            visitDay=data['宿泊数']
            numberOfPeople=data['人数']
            morningType=data['朝食バイキング']
            plan=data['プラン']
            name=data['名前']
            remark=data['備考']

            try:
                super().sendTextAndEnterWaitDisplay(RESERVE_DATE,day)

                super().selectPullDownWaitDisplay(REST_DAY,visitDay)

                super().selectPullDownWaitDisplay(NUMBER_PEOPLE,numberOfPeople)

                super().webElementClick(NEED_MORNING.format(MORNING_TYPE[morningType]))
                if bfPlan!='':
                    # 戻ってきたときチェックボックスが解除されていないので一度解除する
                    super().webElementClick(bfPlan)
                bfPlan=PLAN_SELECT.format(PLAN_TYPE[plan])

                super().webElementClick(PLAN_SELECT.format(PLAN_TYPE[plan]))
                super().sendText(NAME,name)
                super().getScreenShot(screenShotName=name+'_入力画面')
                super().webElementClick(NEXT_BUTTON)

                # 次の画面に遷移後
                # super().waitWebElementVisibility(PAY_DETAIL)
                super().getScreenShot(screenShotName=name+'_確認画面')
                # super().webElementClick(PAY_DETAIL)
                # アラートはスクリーンショットがとれないのでコメントアウト
                # super().getScreenShot(screenShotName=name+'_確認画面_料金詳細')
                # driver.switch_to.alert.accept()
                super().webElementClick(CONFIRM_BUTTON)

                # 完了画面
                super().getScreenShot(screenShotName=name+'_完了画面',sleepTime=5)

                # 初期画面に戻る
                super().webElementClickWaitDisplay(RESULT_BACK_BUTTON)

                # 画面の状態を取得する
                key=str(index+1)+'_'+name
                confirmCost=super().getWebElementTextWaitDisplay(CONFIRM_COST)
                confirmDateFrom=super().getWebElementTextWaitDisplay(CONFIRM_DATE_FROM)
                confirmDateTo=super().getWebElementTextWaitDisplay(CONFIRM_DATE_TO)
                confirmDaycount=super().getWebElementTextWaitDisplay(CONFIRM_DAYSCOUNT)
                confirmPoopleCount=super().getWebElementTextWaitDisplay(CONFIRM_PEOPLE_COUNT)
                confirmNeedMorning=super().getWebElementTextWaitDisplay(CONFIRM_NEED_MORNING)
                confirmPlan=super().getWebElementTextWaitDisplay(CONFIRM_PLAN)
                confirmName=super().getWebElementTextWaitDisplay(CONFIRM_NAME)
                infos=[confirmCost,confirmDateFrom,confirmDateTo,confirmDaycount,confirmPoopleCount,confirmNeedMorning,confirmPlan,confirmName]

                restInfos[key]=infos

                super().webElementClickWaitDisplay(BACK_BUTTON)
            except TimeoutError :
                super().log.error('画面構成が想定外のため失敗:')
                super().log.error(data)
                super().getScreenShot(screenShotName='ErrorInfo',sleepTime=2)
            except TimeoutError :
                super().log.error('登録失敗:')
                super().log.error(data)
                super().getScreenShot(screenShotName='ErrorInfo',sleepTime=2)

        fileName='宿泊情報'+datetime.now().strftime("%Y%m%d%H%M")
        if len(restInfos)==0:
            # 結果がないときは空の配列を返す
            return []
        csvFilePath=self.outPutInfoCSV(iniFile,COLUMN_INFO,restInfos,csvFileName=fileName)
        excelFilePath=self.outPutInfoExcel(iniFile,COLUMN_INFO,COLUMN_SIZE_INFO,restInfos,csvFileName=fileName)
        resultInfo=[csvFilePath,excelFilePath]
        return resultInfo

    # 画面の状態を取得して出力するためメソッドを移植
    def outPutInfoCSV(self,iniFile,colums,infoDict,csvFileName='OutPutInfo',targetDateHourMinute=datetime.now().strftime("%Y%m%d%H%M")):
        infoDirectry = iniFile.get(
            "files", "infoDirectry") + targetDateHourMinute
        os.makedirs(infoDirectry, exist_ok=True)

        infoDataFrame=pd.DataFrame(columns=colums)
        indexCount=1
        for key,dataList in infoDict.items():
            # キーの情報、各種情報のリストの形でまとめる

            infoList=[str(indexCount),key]
            infoList.extend(dataList)
            infoSeries=pd.Series(infoList,index=colums,name=indexCount)
            indexCount=indexCount+1
            infoDataFrame=infoDataFrame.append(infoSeries)
        
        errorOutputFileName=infoDirectry+'/{0}.csv'.format(csvFileName)

        infoDataFrame.to_csv(errorOutputFileName,encoding='cp932',index=False)
        return errorOutputFileName

    # 画面の状態を取得して出力するためメソッドを移植
    def outPutInfoExcel(self,iniFile,colums,columSize,infoDict,csvFileName='OutPutInfo',targetDateHourMinute=datetime.now().strftime("%Y%m%d%H%M")):
        infoDirectry = iniFile.get(
            "files", "infoDirectry") + targetDateHourMinute
        os.makedirs(infoDirectry, exist_ok=True)
        infoDataFrame=pd.DataFrame(columns=colums)
        indexCount=1
        for key,dataList in infoDict.items():
            # キーの情報、各種情報のリストの形でまとめる
            infoList=[str(indexCount),key]
            infoList.extend(dataList)
            infoSeries=pd.Series(infoList,index=colums,name=indexCount)
            indexCount=indexCount+1
            infoDataFrame=infoDataFrame.append(infoSeries)
        
        outputFileName=infoDirectry+'/{0}.xlsx'.format(csvFileName)

        # 結果を一度エクセルに出力
        infoDataFrame.to_excel(outputFileName,encoding='cp932',index=False)

        # 罫線を引くようの設定
        border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'))

        # ヘッダーの色付け
        fill = excel.styles.PatternFill(patternType='solid',
                                        fgColor='87ceeb', bgColor='87ceeb')


        # フォーマットを整える
        targetBook=excel.load_workbook(outputFileName)
        sheet=targetBook.active
        columLength=len(colums)
        
        for index in range(0,columLength):
            # 呼び出し元でカラムのサイズを指定して渡す
            sheet.column_dimensions[get_column_letter(index+1)].width=columSize[index]

        for index,row in enumerate(sheet.rows):
            for cell in row:
                cell.border=border
                if index==0:
                    cell.fill=fill

        targetBook.save(filename=outputFileName)
        return outputFileName



class SeleniumTestSite(WebExecuteBase):
    inputFile=''
    def init(self,inputFile,mode=0,filePaths=['resources/appConfig.ini']):
        super().init(iniFile,mode=mode,filePaths=filePaths)
        self.inputFile=inputFile
    
    def mainExecute(self):
        TARGET_URL=iniFile.get('info','url2')
        log=LoggerObj()
        driver.get(TARGET_URL)
        excelFile=pd.ExcelFile(self.inputFile)
        reserveSheetTemp=excelFile.parse(sheet_name='予約シート',dtype='str',header=1)
        print(reserveSheetTemp.head())

        reserveSheet=reserveSheetTemp.query('無効フラグ != "1"')

        testSideOrder=TestSiteOrder(driver,log,'test')
        # 勤務時間入力
        infos=testSideOrder.inputOrder(reserveSheet)
        # 結果を格納
        super().setResultPath(infos)
        time.sleep(2)

        testSideOrder.createOkDialog('処理完了','登録処理完了')

        #driver.close()


        

# メイン処理
if __name__=="__main__":
    seleniumTestSite =SeleniumTestSite()
    seleniumTestSite.init('data/予約データ2.xlsx')
    seleniumTestSite.mainExecute()



